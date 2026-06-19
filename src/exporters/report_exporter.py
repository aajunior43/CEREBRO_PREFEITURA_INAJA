# src/exporters/report_exporter.py
import csv
import openpyxl
from io import BytesIO, StringIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas

def fmt_brl(v: float) -> str:
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def fmt_brl_signed(v: float) -> str:
    sign = "+" if v >= 0 else ""
    return f"{sign}{fmt_brl(v)}"

# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────
def build_excel(transactions: list, investments: list, alerts: list, stats: dict) -> bytes:
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Resumo Executivo ────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Executivo"
    
    header_fill = PatternFill(start_color="1A1A28", fill_type="solid")
    header_font = Font(bold=True, color="A5B4FC", size=11)
    title_font = Font(bold=True, size=14, color="6366F1")
    bold_font = Font(bold=True)
    
    ws_resumo.cell(row=1, column=1, value="ExpertMoney Analyzer — Relatório de Análise").font = title_font
    ws_resumo.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    resumo_data = [
        [],
        ["INDICADOR", "VALOR"],
        ["Total de Transações", stats.get("totalTransactions", 0)],
        ["Total Entradas (R$)", stats.get("totalCredits", 0.0)],
        ["Total Saídas (R$)", stats.get("totalDebits", 0.0)],
        ["Total Aplicado (R$)", stats.get("totalAplic", 0.0)],
        ["Total Resgatado (R$)", stats.get("totalResgate", 0.0)],
        ["Rendimento Acumulado (R$)", stats.get("totalRendimento", 0.0)],
        ["Saldo Atual na Aplicação (R$)", stats.get("lastBalance", 0.0)],
        ["Último Período", stats.get("lastPeriod", "")],
        [],
        ["ALERTAS", ""],
        ["Total de Alertas", stats.get("totalAlerts", 0)],
        ["Críticos 🔴", stats.get("criticalAlerts", 0)],
        ["Atenção 🟡", stats.get("warningAlerts", 0)],
        ["Informativos 🔵", stats.get("infoAlerts", 0)],
        ["Transações Sinalizadas", stats.get("flaggedCount", 0)],
    ]
    
    for r_idx, row_data in enumerate(resumo_data, start=4):
        if not row_data:
            continue
        ws_resumo.cell(row=r_idx, column=1, value=row_data[0])
        ws_resumo.cell(row=r_idx, column=2, value=row_data[1])
        if row_data[0] in ("INDICADOR", "ALERTAS"):
            ws_resumo.cell(row=r_idx, column=1).font = header_font
            ws_resumo.cell(row=r_idx, column=1).fill = header_fill
            ws_resumo.cell(row=r_idx, column=2).fill = header_fill
            
    ws_resumo.column_dimensions["A"].width = 30
    ws_resumo.column_dimensions["B"].width = 20

    # ── Sheet 2: Transações ──────────────────────────────────
    ws_tx = wb.create_sheet("Transações")
    tx_headers = ["Data", "Tipo", "Valor (R$)", "Direção", "Descrição", "Período", "Arquivo", "Sinalizado", "Alertas"]
    for col_idx, h in enumerate(tx_headers, start=1):
        cell = ws_tx.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        
    sorted_tx = sorted(transactions, key=lambda x: x.get("date") or "")
    for row_idx, t in enumerate(sorted_tx, start=2):
        ws_tx.cell(row=row_idx, column=1, value=t.get("dateStr", ""))
        ws_tx.cell(row=row_idx, column=2, value=t.get("type", ""))
        ws_tx.cell(row=row_idx, column=3, value=t.get("amount", 0.0))
        ws_tx.cell(row=row_idx, column=4, value=t.get("direction", ""))
        ws_tx.cell(row=row_idx, column=5, value=t.get("memo", ""))
        ws_tx.cell(row=row_idx, column=6, value=t.get("period", ""))
        ws_tx.cell(row=row_idx, column=7, value=t.get("filename", ""))
        ws_tx.cell(row=row_idx, column=8, value="SIM" if t.get("flagged") else "")
        ws_tx.cell(row=row_idx, column=9, value=", ".join(t.get("flags", [])))
        
    ws_tx.column_dimensions["A"].width = 12
    ws_tx.column_dimensions["B"].width = 8
    ws_tx.column_dimensions["C"].width = 14
    ws_tx.column_dimensions["D"].width = 8
    ws_tx.column_dimensions["E"].width = 70
    ws_tx.column_dimensions["F"].width = 10
    ws_tx.column_dimensions["G"].width = 35
    ws_tx.column_dimensions["H"].width = 10
    ws_tx.column_dimensions["I"].width = 20

    # ── Sheet 3: Divergências ────────────────────────────────
    ws_div = wb.create_sheet("Divergências")
    div_headers = ["Severidade", "Categoria", "Título", "Descrição", "Evidências"]
    for col_idx, h in enumerate(div_headers, start=1):
        cell = ws_div.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        
    for row_idx, a in enumerate(alerts, start=2):
        evidence_str = " | ".join(f"{e.get('label', '')}: {e.get('value', '')}" for e in a.get("evidence", []))
        ws_div.cell(row=row_idx, column=1, value=str(a.get("severity", "")).upper())
        ws_div.cell(row=row_idx, column=2, value=a.get("category", ""))
        ws_div.cell(row=row_idx, column=3, value=a.get("title", ""))
        ws_div.cell(row=row_idx, column=4, value=a.get("description", ""))
        ws_div.cell(row=row_idx, column=5, value=evidence_str)
        
    ws_div.column_dimensions["A"].width = 12
    ws_div.column_dimensions["B"].width = 35
    ws_div.column_dimensions["C"].width = 60
    ws_div.column_dimensions["D"].width = 80
    ws_div.column_dimensions["E"].width = 100

    # ── Sheet 4: Extrato da Aplicação ────────────────────────
    ws_inv = wb.create_sheet("Aplicação Financeira")
    inv_headers = ["Período", "Saldo Anterior (R$)", "Aplicações (R$)", "Resgates (R$)", "Rend. Bruto (R$)", "IR (R$)", "Saldo Atual (R$)", "Rent. Mês (%)", "Rent. 12m (%)"]
    for col_idx, h in enumerate(inv_headers, start=1):
        cell = ws_inv.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        
    monthly_invs = stats.get("monthlyInvestments", [])
    for row_idx, i in enumerate(monthly_invs, start=2):
        ws_inv.cell(row=row_idx, column=1, value=i.get("period", ""))
        ws_inv.cell(row=row_idx, column=2, value=i.get("saldoAnterior", 0.0))
        ws_inv.cell(row=row_idx, column=3, value=i.get("aplicacoes", 0.0))
        ws_inv.cell(row=row_idx, column=4, value=i.get("resgates", 0.0))
        ws_inv.cell(row=row_idx, column=5, value=i.get("rendBruto", 0.0))
        ws_inv.cell(row=row_idx, column=6, value=i.get("ir", 0.0))
        ws_inv.cell(row=row_idx, column=7, value=i.get("saldoAtual", 0.0))
        ws_inv.cell(row=row_idx, column=8, value=i.get("rentMonth", 0.0))
        ws_inv.cell(row=row_idx, column=9, value=i.get("rentY12", 0.0))
        
    for col_idx in range(1, 10):
        ws_inv.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

# ─────────────────────────────────────────────────────────────
# PDF EXPORT (ReportLab with Custom Canvas for Page Numbers)
# ─────────────────────────────────────────────────────────────
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 7)
        self.setFillColor(colors.HexColor("#64748b"))
        text = f"ExpertMoney Analyzer  |  Página {self._pageNumber} de {page_count}  |  Confidencial"
        self.drawCentredString(A4[0] / 2.0, 30, text)
        self.restoreState()

def build_pdf(transactions: list, investments: list, alerts: list, stats: dict, account: dict = None) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=20*mm,
    )
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        "PDFTitle", parent=styles["Title"],
        fontSize=20, textColor=colors.white,
        alignment=TA_LEFT, fontName="Helvetica-Bold"
    )
    subtitle_style = ParagraphStyle(
        "PDFSubtitle", parent=styles["Normal"],
        fontSize=11, textColor=colors.HexColor("#cbd5e1"),
        fontName="Helvetica"
    )
    section_style = ParagraphStyle(
        "PDFSectionHeading", parent=styles["Heading2"],
        fontSize=12, textColor=colors.HexColor("#1e293b"),
        fontName="Helvetica-Bold", spaceBefore=15, spaceAfter=8
    )
    text_normal = ParagraphStyle(
        "PDFNormalText", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#334155"),
        fontName="Helvetica", leading=11
    )
    text_bold = ParagraphStyle(
        "PDFBoldText", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#0f172a"),
        fontName="Helvetica-Bold", leading=11
    )
    text_muted = ParagraphStyle(
        "PDFMutedText", parent=styles["Normal"],
        fontSize=7.5, textColor=colors.HexColor("#64748b"),
        fontName="Helvetica", leading=10
    )
    
    story = []
    
    # ── Header Banner on Cover ──────────────────────────────
    banner_data = [
        [Paragraph("ExpertMoney Analyzer", title_style)],
        [Paragraph("Relatório de Análise Financeira", subtitle_style)],
        [Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}   |   Conta: {account.get('number', '—') if account else '—'} ({account.get('bank', '—') if account else '—'})", subtitle_style)]
    ]
    banner_table = Table(banner_data, colWidths=[doc.width])
    banner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#6366f1")),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING", (0, 0), (-1, -1), 15),
        ("RIGHTPADDING", (0, 0), (-1, -1), 15),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 15))

    # ── KPI Cards ──────────────────────────────────────────
    story.append(Paragraph("Resumo Geral", section_style))
    kpi_cols = [doc.width / 4.0] * 4
    kpis_data = [
        [
            Paragraph("<b>TRANSAÇÕES</b>", text_muted),
            Paragraph("<b>ALERTAS CRÍTICOS</b>", text_muted),
            Paragraph("<b>RENDIMENTO TOTAL</b>", text_muted),
            Paragraph("<b>SALDO APLICAÇÃO</b>", text_muted)
        ],
        [
            Paragraph(f"<font size=13 color='#6366f1'><b>{stats.get('totalTransactions', 0)}</b></font>", text_bold),
            Paragraph(f"<font size=13 color='#ef4444'><b>{stats.get('criticalAlerts', 0)}</b></font>", text_bold),
            Paragraph(f"<font size=11 color='#6366f1'><b>R$ {fmt_brl(stats.get('totalRendimento', 0.0))}</b></font>", text_bold),
            Paragraph(f"<font size=11 color='#6366f1'><b>R$ {fmt_brl(stats.get('lastBalance', 0.0))}</b></font>", text_bold)
        ]
    ]
    kpis_table = Table(kpis_data, colWidths=kpi_cols)
    kpis_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(kpis_table)
    story.append(Spacer(1, 15))

    # ── Divergências ───────────────────────────────────────
    story.append(Paragraph("Divergências Detectadas", section_style))
    if alerts:
        for idx, a in enumerate(alerts):
            severity = a.get("severity", "info")
            bar_color = colors.HexColor("#ef4444") if severity == "critical" else colors.HexColor("#f59e0b") if severity == "warning" else colors.HexColor("#06b6d4")
            
            alert_header = Paragraph(f"<b><font color='{bar_color.hexval()}'>{a.get('category', '').upper()}</font></b> — {a.get('title', '')}", text_bold)
            alert_body = Paragraph(a.get("description", ""), text_normal)
            
            # Evidências
            ev_list = []
            for ev in a.get("evidence", []):
                ev_list.append(f"<b>{ev.get('label')}:</b> {ev.get('value')}")
            alert_ev = Paragraph(", ".join(ev_list), text_muted) if ev_list else Paragraph("", text_normal)

            alert_table_data = [
                [alert_header],
                [alert_body],
                [alert_ev]
            ]
            alert_table = Table(alert_table_data, colWidths=[doc.width - 6])
            alert_table.setStyle(TableStyle([
                ("LINELEFT", (0, 0), (0, -1), 3, bar_color),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(alert_table)
            story.append(Spacer(1, 8))
    else:
        no_alert_table = Table([[Paragraph("<font color='#10b981'><b>✓ Nenhuma divergência detectada</b></font>", text_bold)]], colWidths=[doc.width])
        no_alert_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0fdf4")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#10b981")),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER")
        ]))
        story.append(no_alert_table)
    
    story.append(Spacer(1, 15))

    # ── Extrato da Aplicação Table ──────────────────────────
    monthly_invs = stats.get("monthlyInvestments", [])
    if monthly_invs:
        story.append(Paragraph("Extrato da Aplicação — BB RF Curto Prazo", section_style))
        inv_w = [18*mm, 24*mm, 24*mm, 24*mm, 24*mm, 26*mm, 18*mm]
        
        inv_table_data = [[
            Paragraph("<b>Período</b>", text_bold),
            Paragraph("<b>Saldo Ant.</b>", text_bold),
            Paragraph("<b>Aplicações</b>", text_bold),
            Paragraph("<b>Resgates</b>", text_bold),
            Paragraph("<b>Rendimento</b>", text_bold),
            Paragraph("<b>Saldo Atual</b>", text_bold),
            Paragraph("<b>Rent.Mês%</b>", text_bold)
        ]]
        
        for inv in monthly_invs:
            rent_val = f"{inv.get('rentMonth', 0.0):.4f}%" if inv.get('rentMonth') is not None else "-"
            inv_table_data.append([
                Paragraph(inv.get("period", ""), text_normal),
                Paragraph(fmt_brl(inv.get("saldoAnterior", 0.0)), text_normal),
                Paragraph(fmt_brl(inv.get("aplicacoes", 0.0)), text_normal),
                Paragraph(fmt_brl(inv.get("resgates", 0.0)), text_normal),
                Paragraph(fmt_brl(inv.get("rendBruto", 0.0)), text_normal),
                Paragraph(fmt_brl(inv.get("saldoAtual", 0.0)), text_normal),
                Paragraph(rent_val, text_normal)
            ])
            
        inv_table = Table(inv_table_data, colWidths=inv_w)
        inv_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")])
        ]))
        # Modify header text color programmatically since reportlab Table needs white color on header Paragraphs
        for c in range(len(inv_w)):
            inv_table_data[0][c].style.textColor = colors.white
        story.append(inv_table)

    story.append(Spacer(1, 15))

    # ── Últimas Transações ──────────────────────────────────
    if transactions:
        story.append(Paragraph(f"Últimas Transações ({min(len(transactions), 30)} de {len(transactions)})", section_style))
        tx_w = [20*mm, 18*mm, 28*mm, doc.width - 66*mm]
        
        tx_table_data = [[
            Paragraph("<b>Data</b>", text_bold),
            Paragraph("<b>Tipo</b>", text_bold),
            Paragraph("<b>Valor (R$)</b>", text_bold),
            Paragraph("<b>Descrição</b>", text_bold)
        ]]
        # Header text color to white
        for c in range(len(tx_w)):
            tx_table_data[0][c].style.textColor = colors.white

        recent_txs = sorted(transactions, key=lambda x: x.get("date") or "", reverse=True)[:30]
        for idx, t in enumerate(recent_txs):
            is_flg = t.get("flagged")
            row_style = text_bold if is_flg else text_normal
            if is_flg:
                row_style = ParagraphStyle("FlgStyle", parent=row_style, textColor=colors.HexColor("#ef4444"))
            
            tx_table_data.append([
                Paragraph(t.get("dateStr", ""), row_style),
                Paragraph(t.get("type", ""), row_style),
                Paragraph(fmt_brl_signed(t.get("amount", 0.0)), row_style),
                Paragraph(t.get("memo", "")[:55], row_style)
            ])
            
        tx_table = Table(tx_table_data, colWidths=tx_w)
        tx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")])
        ]))
        
        # Apply specific row backgrounds for flagged items
        for idx, t in enumerate(recent_txs, start=1):
            if t.get("flagged"):
                tx_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#fef2f2"))
                ]))
        story.append(tx_table)

    story.append(Spacer(1, 15))

    # ── Apêndice: Transações Sinalizadas ────────────────────
    flagged_txs = sorted([t for t in transactions if t.get("flagged")], key=lambda x: x.get("date") or "")
    if flagged_txs:
        story.append(PageBreak())
        story.append(Paragraph(f"Apêndice — Transações Sinalizadas ({len(flagged_txs)} no total)", section_style))
        story.append(Paragraph("Todas as transações que geraram pelo menos um alerta nesta análise.", text_muted))
        story.append(Spacer(1, 6))
        
        apx_w = [18*mm, 15*mm, 24*mm, 28*mm, doc.width - 85*mm]
        apx_table_data = [[
            Paragraph("<b>Data</b>", text_bold),
            Paragraph("<b>Tipo</b>", text_bold),
            Paragraph("<b>Valor (R$)</b>", text_bold),
            Paragraph("<b>Flags</b>", text_bold),
            Paragraph("<b>Descrição</b>", text_bold)
        ]]
        for c in range(len(apx_w)):
            apx_table_data[0][c].style.textColor = colors.white
            
        for t in flagged_txs:
            row_style = ParagraphStyle("ApxRowStyle", parent=text_normal, textColor=colors.HexColor("#ef4444"))
            flag_str = ", ".join(t.get("flags", []))[:20] or "—"
            apx_table_data.append([
                Paragraph(t.get("dateStr", ""), row_style),
                Paragraph(t.get("type", ""), row_style),
                Paragraph(fmt_brl_signed(t.get("amount", 0.0)), row_style),
                Paragraph(flag_str, row_style),
                Paragraph(t.get("memo", "")[:45], row_style)
            ])
            
        apx_table = Table(apx_table_data, colWidths=apx_w)
        apx_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ef4444")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#fef2f2"))
        ]))
        story.append(apx_table)

    doc.build(story, canvasmaker=NumberedCanvas)
    return buffer.getvalue()

# ─────────────────────────────────────────────────────────────
# CSV EXPORT
# ─────────────────────────────────────────────────────────────
def generate_csv_transactions(transactions: list) -> str:
    output = StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_NONNUMERIC)
    writer.writerow(["Data", "Tipo", "Valor", "Direção", "Memo", "Período", "Alertado", "Flags"])
    for tx in transactions:
        writer.writerow([
            tx.get("dateStr", ""),
            tx.get("type", ""),
            str(abs(tx.get("amount", 0.0))).replace(".", ","),
            "Débito" if tx.get("amount", 0.0) < 0 else "Crédito",
            tx.get("memo", ""),
            tx.get("period", ""),
            "Sim" if tx.get("flagged") else "Não",
            "|".join(tx.get("flags", []))
        ])
    return output.getvalue()

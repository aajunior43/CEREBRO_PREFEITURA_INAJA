"""
exportar_dados.py — Exporta credores do Excel para data.js
Prefeitura Municipal de Inajá

USO:
  python exportar_dados.py

Isso irá ler o arquivo CREDORES_mes2_destacado.xlsx e recriar o data.js.
Execute sempre que atualizar o arquivo Excel.
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f'{sys.executable} -m pip install openpyxl')
    import openpyxl

EXCEL_FILE = 'CREDORES_mes2_destacado.xlsx'
OUTPUT_FILE = 'data.js'

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f'ERRO: Arquivo "{EXCEL_FILE}" não encontrado.')
        print('Certifique-se de que o Excel está na mesma pasta que este script.')
        input('Pressione ENTER para sair...')
        return

    print(f'Lendo {EXCEL_FILE}...')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    data = []
    uid = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and v is not False for v in row):
            continue
        row_dict = {'id': uid}
        for h, v in zip(headers, row):
            key = str(h) if h is not None else 'OBS'
            if key == 'None':
                key = 'OBS'
            if isinstance(v, (int, float)):
                row_dict[key] = float(v)
            elif v is None or v is False:
                row_dict[key] = ''
            else:
                row_dict[key] = str(v).strip()
        # Padroniza campos ausentes
        row_dict.setdefault('OBS', '')
        row_dict.setdefault('DEPARTAMENTO', '')
        data.append(row_dict)
        uid += 1

    # Gera o JS
    js_content = (
        '// Dados dos credores fixos mensais - Prefeitura Municipal de Inajá\n'
        '// Gerado automaticamente por exportar_dados.py\n'
        '// Para atualizar: python exportar_dados.py\n\n'
        f'const CREDORES_FIXOS = {json.dumps(data, ensure_ascii=False, indent=2)};\n'
    )

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f'✓ {len(data)} credores exportados para {OUTPUT_FILE}')
    print('Pronto! Abra o index.html no navegador para ver as atualizações.')
    input('Pressione ENTER para sair...')

if __name__ == '__main__':
    main()
